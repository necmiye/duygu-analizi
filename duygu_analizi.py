import json
import urllib.request
from  openpyxl import *
import numpy as np
import cv2


req = urllib.request.Request('http://betikdiller.ylmz.pw/index?count=20')
with urllib.request.urlopen(req) as response:
    result = json.loads(response.read().decode('utf-8'))

Karamsar = ["imkansız","intihar","depresyon","gözyaşı","kötü","bunalım","başarısızlık","nefret","yalnızlık","yoruldum","iyilesemiyor","halledebilirdi","kırıldım","yanlış","inanmıstım","hissetmedim","sıkıntı"]
Yaratıcı = ["merak","sanat","yenilik","icat","ilham","bakıs","acısı","açısı","proje","yaratıcı","yaratici","yenilige","yeniliğe","acık","açık","acik","şiir","sanatsal","olagan","olağan","üstü"]
iyimser = ["mutlu","yardımsever","başarı","pozitif","sevgi","umut","güzel",":)","poncik","ponçik","mutluyum","kendim","seviyorum","gülümse","kahkaha","degerli","değerli","umutluyum","pozitif"]
Mizahi = ["espri","şaka","eğlence","komik","komedi","kariketür ","caps","mood","kanka","hahaha","haha","karikatur","karikatür","kral","nasıl","sen","ben"]
Dısa_donuk = ["altivite","dolasırım","dolasirim","dolaşırım","gez","gezmeli","tatil","kamp","camp","parti","partile","seyehat","özgürlük","ozgurluk","özgürce","etkinlik","etki","faal","faaliyet","karavan","doğa","doga","yasam","yaşam","yaşama","sevinci"]
karakterler = [Karamsar,Yaratıcı,iyimser,Mizahi,Dısa_donuk]
sayac = 0
kisi = [0]

kitap = load_workbook("Kitap1.xlsx")
sheet = kitap.active
for i in range(len(result)):
    sozluk = dict(result[i])
    aranan = sozluk["text"]
    
    for j in range(len(karakterler)):
        a = karakterler[j]
        for x in range(len(a)):
            if a[x] in aranan:
                sayac = sayac + 1
        if(sayac>0):
            if ( karakterler[j][0]== Karamsar[0]):
                kisi.append(1)
            elif ( karakterler[j][0]== Yaratıcı[0]):
                kisi.append(2)
            elif ( karakterler[j][0]== iyimser[0]):
                kisi.append(3)
            elif ( karakterler[j][0]== Mizahi[0]):
                kisi.append(4)
            elif ( karakterler[j][0]== Dısa_donuk[0]):
                kisi.append(5)
            sayac=0
    
    sheet.append((sozluk["id"],sozluk["text"]))
    kitap.save("Kitap1.xlsx")
kitap.close()

Karamsar1 = 0.0
Yaratıcı1 = 0.0
iyimser1 = 0.0
Mizahi1 = 0.0
Dısa_donuk1 = 0.0

for z in range(len(kisi)):
    toplam = kisi.count(1) + kisi.count(2) +kisi.count(3) + kisi.count(4)+ kisi.count(5)
    if (toplam == 0.0):
        toplam = 1.0
    Karamsar1 = kisi.count(1)/toplam
    Yaratıcı1 = kisi.count(2)/toplam
    iyimser1 = kisi.count(3)/toplam
    Mizahi1 = kisi.count(4)/toplam
    Dısa_donuk1 = kisi.count(5)/toplam

oran_listesi = [Karamsar1,Yaratıcı1,iyimser1,Mizahi1,Dısa_donuk1]
en_yuksek = 0.0
for i in range(len(oran_listesi)):
    print ( oran_listesi[i])
    if (en_yuksek < oran_listesi[i]):
        en_yuksek = oran_listesi[i]
        print(en_yuksek)
if (en_yuksek== iyimser1):
    resim = cv2.imread('iyimser1.jpeg')
    cv2.imshow('iyimser karakter',resim)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
elif(en_yuksek==Karamsar1):
    resim = cv2.imread('karamsar1.jfif')
    cv2.imshow('karamsar karakter',resim)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
elif(en_yuksek==Yaratıcı1):
    resim = cv2.imread('yaratici1.jpeg')
    cv2.imshow('yaratici karakter',resim)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
elif(en_yuksek==Mizahi1):
    resim = cv2.imread('mizahi1.jpeg')
    cv2.imshow('mizahi karakter',resim)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
else:
    resim = cv2.imread('disa_donuk.jpeg')
    cv2.imshow('dışa dönük karakter',resim)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
print ("kisi " + str(Karamsar1) + " oranında karamsar")
print ("kisi " + str(Yaratıcı1) + " oranında yaratıcı")
print ("kisi " + str(iyimser1) + " oranında iyimser")
print ("kisi " + str(Mizahi1) + " oranında mizahi")
print ("kisi " + str(Dısa_donuk1) + " oranında dısa donuk")

'''
EXELDEN OKUMAK VERİ İÇİN 
kitap = load_workbook("Kitap1.xlsx")
sheet=kitap.active
for row in sheet.iter_rows(min_row=1, min_col=1, max_row=4, max_col=2):
    for cell in row:
        print(cell.value ,end=" ")
    print()

kitap.close()'''